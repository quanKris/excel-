import pandas as pd
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import copy

# 读取第二个 Excel 文件
file_path = 'C:\\Users\\MDPI\\Desktop\\脚本\\ddd.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active # 假设数据在第一个 sheet 中

# 打开模板文件
template_file_path = 'C:\\Users\\MDPI\\Desktop\\脚本\\ccc.xlsx'
template_workbook = load_workbook(template_file_path, data_only=True)
template_sheet = template_workbook.active




# 循环处理每一行数据
for row_index in range(3, sheet.max_row + 1):  # 从第三行开始到最后一行
    # 创建新的 Workbook 和 Worksheet
    new_wb = Workbook()
    new_ws = new_wb.active
    
    # 获取行数据
    row_data = [cell.value for cell in sheet[row_index]]
    
    # 获取第二个和第五个值
    
    i = 1 
    # 复制模板文件到新的 Excel 文件中
    for row in template_sheet.iter_rows():
        for cell in row:
            template_cell = template_sheet[cell.coordinate]
            new_cell = new_ws[cell.coordinate]
            new_cell.value = template_cell.value
            
            # 复制模板的样式
            new_cell.font = copy.copy(template_cell.font)
            new_cell.fill = copy.copy(template_cell.fill)
            new_cell.border = copy.copy(template_cell.border)
            new_cell.alignment = copy.copy(template_cell.alignment)
            new_cell.number_format = template_cell.number_format
            # # 处理合并单元格
            # if template_cell.coordinate in template_sheet.merged_cells:
            #     new_ws.merge_cells(template_cell.coordinate)
    value_2 = row_data[1]  # 假设第二个值
    value_5 = row_data[4]  # 假设第五个值
    # 将值插入到新的 Excel 文件中的指定位置
    new_ws.cell(row=2, column=4, value=value_2)
    new_ws.cell(row=3, column=4, value=value_5)


    # 合并单元格
    merged_range = "A1:N1"
    new_ws.merge_cells(merged_range)
    new_ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=9)
    new_ws.merge_cells(start_row=28, start_column=1, end_row=28, end_column=11)
    new_ws.merge_cells(start_row=31, start_column=1, end_row=31, end_column=14)
    for row_num in range(20, 28):  # 20 to 27
        new_ws.row_dimensions[row_num].height = 20
    new_ws.merge_cells(start_row=20, start_column=1, end_row=27, end_column=14)
    new_ws.merge_cells(start_row=5, start_column=1, end_row=8, end_column=1)
    new_ws.merge_cells(start_row=9, start_column=1, end_row=12, end_column=1)
    new_ws.merge_cells(start_row=13, start_column=1, end_row=14, end_column=1)
    new_ws.merge_cells(start_row=15, start_column=1, end_row=19, end_column=1)
    for row_num in range(5, 20):  # 5 to 19
        new_ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=9)
    for row_num in range(2, 4): 
        new_ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    for row_num in range(2, 4): 
        new_ws.merge_cells(start_row=row_num, start_column=4, end_row=row_num, end_column=9)
    for row_num in range(2, 4): 
        new_ws.merge_cells(start_row=row_num, start_column=10, end_row=row_num, end_column=12)
    for row_num in range(2, 4): 
        new_ws.merge_cells(start_row=row_num, start_column=13, end_row=row_num, end_column=14)
    for row_num in range(4, 20): 
        new_ws.merge_cells(start_row=row_num, start_column=11, end_row=row_num, end_column=12)
    for row_num in range(29, 31): 
        new_ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    for row_num in range(29, 31): 
        new_ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=8)
    for row_num in range(29, 31): 
        new_ws.merge_cells(start_row=row_num, start_column=10, end_row=row_num, end_column=14)
    
    # 高度
    for row_num in range(1, 20):  # 1 to 19
        new_ws.row_dimensions[row_num].height = 24
    for row_num in range(28, 32):  # 1 to 19
        new_ws.row_dimensions[row_num].height = 24
    for row_num in range(20, 27):  # 1 to 19
        new_ws.row_dimensions[row_num].height = 19

    # 列宽
    new_ws.column_dimensions['A'].width = 6 
    new_ws.column_dimensions['B'].width = 6 
    new_ws.column_dimensions['c'].width = 1.3 
    new_ws.column_dimensions['D'].width = 0.01 
    new_ws.column_dimensions['E'].width = 0.01 
    new_ws.column_dimensions['F'].width = 10.38
    new_ws.column_dimensions['G'].width = 7.5
    new_ws.column_dimensions['H'].width = 13
    new_ws.column_dimensions['I'].width = 10 
    new_ws.column_dimensions['J'].width = 8
    new_ws.column_dimensions['K'].width = 3
    new_ws.column_dimensions['L'].width = 5.2 
    new_ws.column_dimensions['N'].width = 5.9

    new_ws.row_dimensions[4].height = 30


    # 保存新的 Excel 文件，以第二个值作为文件名
    new_file_name = f'{value_2}.xlsx'  # 使用第二个值作为文件名
    new_wb.save(new_file_name)

print("文件生成完毕！")